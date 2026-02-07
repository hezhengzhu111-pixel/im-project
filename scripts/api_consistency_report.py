import argparse
import ast
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import requests


METHODS = {"GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"}


@dataclass(frozen=True)
class Endpoint:
    method: str
    path: str
    service: str
    source: str
    detail: str = ""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--base-url", default=os.environ.get("IM_BASE_URL", "http://localhost:8080"))
    p.add_argument("--services", default="auth,user,group,message,file,im")
    p.add_argument("--frontend-services-dir", default=str(Path("frontend") / "src" / "services"))
    p.add_argument("--tests-glob", default="test_module_*.py")
    p.add_argument("--out", default="")
    return p.parse_args()


def now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalize_path(path: str) -> str:
    if not path:
        return "/"
    if not path.startswith("/"):
        path = "/" + path
    path = re.sub(r"//+", "/", path)
    return path


def normalize_template(s: str) -> str:
    s = re.sub(r"\$\{[^}]+\}", "{param}", s)
    s = re.sub(r"\{[^}]+\}", "{param}", s)
    s = re.sub(r"\d+", "{num}", s)
    return s


def map_openapi_path_to_gateway(service: str, internal_path: str) -> Optional[str]:
    p = normalize_path(internal_path)
    if service == "auth":
        return normalize_path("/api/auth" + p)
    if service == "user":
        if p.startswith("/api/"):
            return p
        if p.startswith("/user/") or p == "/user":
            return normalize_path("/api" + p)
        if p.startswith("/test/") or p.startswith("/user/test/"):
            return normalize_path("/api" + p)
        return normalize_path("/api" + p)
    if service == "group":
        if p.startswith("/s/") or p == "/s":
            return normalize_path("/api/group" + p[2:])
        return normalize_path("/api/group" + p)
    if service == "message":
        if p.startswith("/s/") or p == "/s":
            return normalize_path("/api/message" + p[2:])
        return normalize_path("/api/message" + p)
    if service == "file":
        return normalize_path("/api/file" + p)
    if service == "im":
        if p.startswith("/api/"):
            return p
        return normalize_path("/api/im" + p)
    return None


def fetch_openapi(base_url: str, service: str) -> dict:
    url = base_url.rstrip("/") + f"/v3/api-docs/{service}"
    headers = {"X-Gateway-Route": "true", "X-Trace-Id": f"openapi-{service}-{now_ts()}"}
    resp = requests.get(url, headers=headers, timeout=20)
    resp.raise_for_status()
    return resp.json()


def extract_openapi_endpoints(spec: dict, service: str) -> List[Endpoint]:
    endpoints: List[Endpoint] = []
    paths = spec.get("paths") or {}
    for internal_path, methods in paths.items():
        if not isinstance(methods, dict):
            continue
        for method, op in methods.items():
            m = str(method).upper()
            if m not in METHODS:
                continue
            gw_path = map_openapi_path_to_gateway(service, str(internal_path))
            if not gw_path:
                continue
            detail = ""
            if isinstance(op, dict):
                rb = op.get("requestBody") or {}
                content = (rb.get("content") or {}) if isinstance(rb, dict) else {}
                req_ct = ",".join(sorted(content.keys())) if isinstance(content, dict) else ""
                resp_codes = ",".join(sorted((op.get("responses") or {}).keys())) if isinstance(op.get("responses"), dict) else ""
                detail = f"req_ct={req_ct}; resp={resp_codes}"
            endpoints.append(Endpoint(method=m, path=gw_path, service=service, source="openapi", detail=detail))
    return endpoints


TS_CALL_RE = re.compile(
    r"http\.(get|post|put|delete|patch)\s*(?:<[^>]*>)?\s*\(\s*(?P<q>['\"`])(?P<path>.*?)(?P=q)",
    re.IGNORECASE | re.DOTALL,
)


def extract_frontend_endpoints(services_dir: Path) -> List[Endpoint]:
    endpoints: List[Endpoint] = []
    for ts in sorted(services_dir.glob("*.ts")):
        text = ts.read_text(encoding="utf-8", errors="replace")
        for m in TS_CALL_RE.finditer(text):
            method = m.group(1).upper()
            raw_path = m.group("path").strip()
            if not raw_path:
                continue
            if raw_path.startswith("http://") or raw_path.startswith("https://"):
                continue
            p = normalize_path(raw_path)
            gw_path = normalize_path("/api" + p)
            endpoints.append(Endpoint(method=method, path=gw_path, service="frontend", source=str(ts.name)))
    return endpoints


def _str_from_ast(node: ast.AST) -> Optional[str]:
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return node.value
    if isinstance(node, ast.JoinedStr):
        parts: List[str] = []
        for v in node.values:
            if isinstance(v, ast.Constant) and isinstance(v.value, str):
                parts.append(v.value)
            else:
                parts.append("{param}")
        return "".join(parts)
    return None


class _TestVisitor(ast.NodeVisitor):
    def __init__(self, file_name: str) -> None:
        self.file_name = file_name
        self.items: List[Tuple[str, str]] = []

    def visit_Call(self, node: ast.Call) -> None:
        if isinstance(node.func, ast.Attribute) and node.func.attr == "call":
            args = list(node.args or [])
            if len(args) >= 3:
                method = _str_from_ast(args[1])
                path = _str_from_ast(args[2])
                if method and path and method.upper() in METHODS and str(path).startswith("/"):
                    self.items.append((method.upper(), normalize_path(path)))
        self.generic_visit(node)


def extract_test_endpoints(glob_pattern: str) -> List[Endpoint]:
    endpoints: List[Endpoint] = []
    for py in sorted(Path(".").glob(glob_pattern)):
        try:
            tree = ast.parse(py.read_text(encoding="utf-8", errors="replace"))
        except SyntaxError:
            continue
        v = _TestVisitor(py.name)
        v.visit(tree)
        for method, path in v.items:
            endpoints.append(Endpoint(method=method, path=path, service="tests", source=str(py.name)))
    return endpoints


def choose_output_path(out_arg: str) -> Path:
    if out_arg:
        return Path(out_arg)
    return Path("test_reports") / f"api_consistency_{now_ts()}.xlsx"


def write_xlsx(
    out_path: Path,
    rows: Sequence[dict],
    meta: dict,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:
        csv_path = out_path.with_suffix(".csv")
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        header = list(rows[0].keys()) if rows else []
        with csv_path.open("w", encoding="utf-8") as f:
            f.write(",".join(header) + "\n")
            for r in rows:
                f.write(",".join(str(r.get(k, "")).replace(",", " ") for k in header) + "\n")
        raise RuntimeError(f"openpyxl not available ({e}); wrote fallback CSV: {csv_path}") from e

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "对照表"

    header = [
        "method",
        "path",
        "in_openapi",
        "in_frontend",
        "in_tests",
        "openapi_detail",
        "frontend_sources",
        "test_sources",
        "diff",
    ]
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red = PatternFill("solid", fgColor="FFCCCC")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    for r in rows:
        ws.append([r.get(k, "") for k in header])
        rr = ws.max_row
        diff = str(r.get("diff") or "")
        if diff:
            fill = red
        else:
            fill = yellow if (not r.get("in_frontend") or not r.get("in_tests")) else None
        if fill:
            for c in range(1, len(header) + 1):
                ws.cell(row=rr, column=c).fill = fill

    ws_meta = wb.create_sheet("元数据")
    ws_meta.append(["key", "value"])
    for k, v in meta.items():
        ws_meta.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)])

    for col, w in zip("ABCDEFGHI", [10, 50, 12, 12, 10, 28, 30, 30, 22]):
        ws.column_dimensions[col].width = w

    wb.save(out_path)


def main() -> int:
    args = parse_args()
    base_url = args.base_url.rstrip("/")
    services = [s.strip() for s in (args.services or "").split(",") if s.strip()]
    services_dir = Path(args.frontend_services_dir)

    openapi_eps: List[Endpoint] = []
    openapi_errors: Dict[str, str] = {}
    for svc in services:
        try:
            spec = fetch_openapi(base_url, svc)
            openapi_eps.extend(extract_openapi_endpoints(spec, svc))
        except Exception as e:
            openapi_errors[svc] = str(e)

    fe_eps = extract_frontend_endpoints(services_dir) if services_dir.exists() else []
    test_eps = extract_test_endpoints(args.tests_glob)

    by_key: Dict[Tuple[str, str], dict] = {}

    def ensure(method: str, path: str) -> dict:
        key = (method, path)
        if key not in by_key:
            by_key[key] = {
                "method": method,
                "path": path,
                "in_openapi": False,
                "in_frontend": False,
                "in_tests": False,
                "openapi_detail": "",
                "frontend_sources": set(),
                "test_sources": set(),
                "diff": "",
            }
        return by_key[key]

    for ep in openapi_eps:
        row = ensure(ep.method, ep.path)
        row["in_openapi"] = True
        if ep.detail:
            row["openapi_detail"] = ep.detail

    for ep in fe_eps:
        row = ensure(ep.method, ep.path)
        row["in_frontend"] = True
        row["frontend_sources"].add(ep.source)

    for ep in test_eps:
        row = ensure(ep.method, ep.path)
        row["in_tests"] = True
        row["test_sources"].add(ep.source)

    rows: List[dict] = []
    for key, row in sorted(by_key.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        diff_parts = []
        if row["in_frontend"] and not row["in_openapi"]:
            diff_parts.append("frontend_only")
        if row["in_openapi"] and not row["in_frontend"]:
            diff_parts.append("missing_in_frontend")
        if row["in_tests"] and not row["in_openapi"]:
            diff_parts.append("tests_only")
        if row["in_openapi"] and not row["in_tests"]:
            diff_parts.append("missing_in_tests")
        row["diff"] = ";".join(diff_parts)
        row["frontend_sources"] = ",".join(sorted(row["frontend_sources"])) if isinstance(row["frontend_sources"], set) else row["frontend_sources"]
        row["test_sources"] = ",".join(sorted(row["test_sources"])) if isinstance(row["test_sources"], set) else row["test_sources"]
        rows.append(row)

    out_path = choose_output_path(args.out)
    meta = {
        "base_url": base_url,
        "services": services,
        "frontend_services_dir": str(services_dir),
        "tests_glob": args.tests_glob,
        "openapi_errors": openapi_errors,
        "generated_at": datetime.now().isoformat(),
    }
    try:
        write_xlsx(out_path, rows, meta)
    except RuntimeError as e:
        print(str(e))
        return 2

    print(f"OK: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

